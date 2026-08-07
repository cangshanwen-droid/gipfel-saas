"""Excel 导入导出"""
import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from ..database import get_db
from ..auth import get_current_user
from ..models import (
    Region, Company, Contract, ContractItem, ContractType,
    InfrastructureType, FormulaLog, RegionAccount, AccountTransaction
)

router = APIRouter()

TABLE_MAP = {
    "regions": (Region, "区域"),
    "companies": (Company, "公司"),
    "contracts": (Contract, "合同"),
    "contract_items": (ContractItem, "合同明细"),
    "contract_types": (ContractType, "合同类型"),
    "infrastructure_types": (InfrastructureType, "基建类型"),
    "formula_logs": (FormulaLog, "模拟日志"),
    "region_accounts": (RegionAccount, "财务账户"),
    "account_transactions": (AccountTransaction, "交易流水"),
}


@router.get("/export")
def export_excel(db: Session = Depends(get_db), _=Depends(get_current_user)):
    wb = Workbook()
    wb.remove(wb.active)

    for table_name, (model, label) in TABLE_MAP.items():
        rows = db.query(model).all()
        if not rows:
            continue

        ws = wb.create_sheet(title=label)
        # 获取列名（排除 _sa_instance_state 和计算属性）
        first = rows[0]
        columns = [c.name for c in model.__table__.columns]

        # 写表头
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # 写数据
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                val = getattr(row, col_name, None)
                ws.cell(row=row_idx, column=col_idx, value=val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gipfel_export.xlsx"},
    )
